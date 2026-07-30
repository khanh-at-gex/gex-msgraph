"""SpreadsheetML 2003 (XML Spreadsheet) → xlsx conversion.

Some systems — notably SAP exports — write ``.xls`` files that are not OLE2
workbooks at all but plain XML in Microsoft's SpreadsheetML 2003 schema,
usually UTF-16 encoded. Excel and SharePoint open them happily because of the
``<?mso-application progid="Excel.Sheet"?>`` hint, but every pandas engine
rejects them: they sniff magic bytes and see ``\\xfe\\xff`` (UTF-16 BOM)
instead of ``\\xd0\\xcf\\x11\\xe0`` (OLE2).

Rather than build DataFrames directly, this module rebuilds the grid into an
``openpyxl`` workbook and serialises it to real xlsx bytes. pandas then reads
those bytes normally, so every ``read_excel`` kwarg (``header``, ``usecols``,
``dtype``, ``skiprows``, …) and the whole sheet-matching machinery in
``read_excel_many`` keep working untouched.
"""

from __future__ import annotations

import logging
import re
import xml.etree.ElementTree as ET
from typing import Any

logger = logging.getLogger("gex_msgraph")

SS_NS = "urn:schemas-microsoft-com:office:spreadsheet"

_Q_WORKSHEET = f"{{{SS_NS}}}Worksheet"
_Q_TABLE = f"{{{SS_NS}}}Table"
_Q_ROW = f"{{{SS_NS}}}Row"
_Q_CELL = f"{{{SS_NS}}}Cell"
_Q_DATA = f"{{{SS_NS}}}Data"
_A_NAME = f"{{{SS_NS}}}Name"
_A_INDEX = f"{{{SS_NS}}}Index"
_A_TYPE = f"{{{SS_NS}}}Type"
_A_MERGE_ACROSS = f"{{{SS_NS}}}MergeAcross"

# The Workbook root (which carries the namespace declaration) sits within the
# first few hundred bytes; 8 KiB leaves room for a long declaration + PIs.
_SNIFF_BYTES = 8192

# Escape a bare "&". The lookahead preserves the five predefined entities and
# numeric character references — and deliberately nothing else, so a reference
# to a DTD-declared entity is neutralised into literal text. That costs us
# nothing (Excel and SAP never declare entities) and it is what keeps
# ElementTree, which does expand internal entities, immune to billion-laughs
# blowups on downloaded files. See test_entity_expansion_is_neutralised.
_BARE_AMP = re.compile(r"&(?!(?:amp|lt|gt|quot|apos|#\d+|#x[0-9a-fA-F]+);)")
_XML_DECL = re.compile(r"^\s*<\?xml[^>]*\?>")
# CDATA content is literal by definition, so a "&" inside it is already valid
# and must survive the repair above untouched.
_CDATA = re.compile(r"<!\[CDATA\[.*?\]\]>", re.DOTALL)
# Excel rejects these in a sheet name; SAP-style writers do not always comply.
_BAD_TITLE_CHARS = re.compile(r"[\\/*?:\[\]]")


def _decode(data: bytes, *, strict: bool = True) -> str:
    """Decode SpreadsheetML bytes using the BOM, stripping any leading BOM char.

    Decodes strictly by default: silently substituting U+FFFD into an account
    name is a worse failure for a finance pipeline than a loud one. Malformed
    input falls back to replacement so a single bad byte cannot cost the whole
    file, but it says so in the log.
    """
    if data.startswith(b"\xfe\xff"):
        codec = "utf-16-be"
    elif data.startswith(b"\xff\xfe"):
        codec = "utf-16-le"
    elif data.startswith(b"\xef\xbb\xbf"):
        codec = "utf-8-sig"
    else:
        codec = "utf-8"

    try:
        text = data.decode(codec, errors="strict" if strict else "replace")
    except UnicodeDecodeError as e:
        logger.warning(
            "SpreadsheetML: %s decode failed at byte %s (%s); "
            "falling back to replacement characters",
            codec,
            e.start,
            e.reason,
        )
        text = data.decode(codec, errors="replace")

    # The explicit utf-16-be/-le codecs (unlike plain "utf-16") leave the BOM
    # in place as U+FEFF, and a leading U+FEFF makes ElementTree raise.
    return text.lstrip("﻿")


def looks_like_spreadsheetml(data: bytes) -> bool:
    """True if `data` is a SpreadsheetML 2003 document.

    Sniffs content only — never the file extension, since the misleading
    ``.xls`` extension is exactly the problem this module exists for. Requires
    both an XML-ish prefix and the SpreadsheetML namespace so real xlsx/xls
    files fall through to pandas unchanged.
    """
    if not data.startswith((b"\xfe\xff", b"\xff\xfe", b"\xef\xbb\xbf", b"<?xml")):
        return False
    # The namespace is ASCII in the source but UTF-16 on the wire, so it has to
    # be matched after decoding. Non-strict here: the prefix is cut at a fixed
    # offset that can land mid code unit, which is expected, not an anomaly.
    return SS_NS in _decode(data[:_SNIFF_BYTES], strict=False)


def sanitize_sheet_title(name: str) -> str:
    """Map a SpreadsheetML sheet name onto a title xlsx can hold."""
    return _BAD_TITLE_CHARS.sub("_", name)[:31]


def prepare_excel_bytes(
    data: bytes,
    sheet: str | int = 0,
    *,
    sanitize_sheet_name: bool = True,
) -> tuple[bytes, str | int]:
    """Convert SpreadsheetML to xlsx when needed; remap sheet name to match.

    Uses the conversion's actual original-name -> final-title map rather than
    recomputing `sanitize_sheet_title` blindly: two distinct source names can
    sanitize to the same title, in which case openpyxl deduplicates the
    second one (`Q1_A` -> `Q1_A1`). Blind recomputation would resolve both
    requests to `Q1_A` and silently hand back the first sheet's data for a
    request naming the second — recompute only as a fallback for a name that
    was never a worksheet in this file, so pandas still raises its normal
    "sheet not found" for a genuine typo.
    """
    if not looks_like_spreadsheetml(data):
        return data, sheet
    data, name_map = _convert(data)
    if sanitize_sheet_name and isinstance(sheet, str):
        sheet = name_map.get(sheet, sanitize_sheet_title(sheet))
    return data, sheet


def _repair_ampersands(text: str) -> str:
    """Escape bare ``&`` everywhere except inside CDATA sections."""
    out: list[str] = []
    pos = 0
    for m in _CDATA.finditer(text):
        out.append(_BARE_AMP.sub("&amp;", text[pos : m.start()]))
        out.append(m.group(0))
        pos = m.end()
    out.append(_BARE_AMP.sub("&amp;", text[pos:]))
    return "".join(out)


def _cell_value(cell: ET.Element) -> Any:
    """Extract a cell's value, coerced per its ``ss:Type``."""
    data_el = cell.find(_Q_DATA)
    if data_el is None:
        return None

    # itertext() rather than .text: <Data> may wrap rich-text runs (<html:B>),
    # and .text would only return the fragment before the first child.
    text = "".join(data_el.itertext())
    if not text:
        return None

    if data_el.get(_A_TYPE) == "Number":
        try:
            return float(text)
        except ValueError:
            return text
    return text


def to_xlsx_bytes(data: bytes) -> bytes:
    """Convert SpreadsheetML 2003 bytes into real xlsx bytes.

    Only values are carried over — styles, formulas and merge geometry are
    dropped, since the sole purpose is to hand pandas something it can read.
    ``ss:Type="DateTime"`` cells come through as strings.
    """
    return _convert(data)[0]


def _convert(data: bytes) -> tuple[bytes, dict[str, str]]:
    """Do the actual conversion; also returns the original -> final sheet
    title map, since openpyxl may rename a title further to dedupe it after
    `sanitize_sheet_title` has already resolved a collision (see
    `prepare_excel_bytes`)."""
    import io

    from openpyxl import Workbook

    text = _decode(data)
    # We hand ElementTree a str, so a declared encoding="UTF-16" would only
    # confuse it — the declaration has done its job already.
    text = _XML_DECL.sub("", text, count=1)
    text = _repair_ampersands(text)

    root = ET.fromstring(text)

    wb = Workbook()
    wb.remove(wb.active)
    name_map: dict[str, str] = {}

    for sheet_idx, worksheet in enumerate(root.findall(_Q_WORKSHEET)):
        name = worksheet.get(_A_NAME) or f"Sheet{sheet_idx + 1}"
        # openpyxl raises on titles Excel would reject. The writers that emit
        # SpreadsheetML do not always respect those rules (this whole module
        # exists because SAP emits unescaped "&"), and a bad sheet name is no
        # reason to lose the file's data.
        safe = sanitize_sheet_title(name) or f"Sheet{sheet_idx + 1}"
        if safe != name:
            logger.warning(
                "SpreadsheetML: sheet name %r is not a legal xlsx title, using %r",
                name,
                safe,
            )
        ws = wb.create_sheet(title=safe)
        # ws.title, not `safe`: openpyxl may append a further suffix here to
        # dedupe against an earlier sheet whose name sanitized to the same
        # string. Recording the real title is what lets a request for this
        # sheet's original name resolve to the right one instead of silently
        # colliding with the first. setdefault, not assignment: if the
        # source itself repeats a worksheet name (already malformed — Excel
        # would refuse to save that), the first occurrence wins, matching
        # what a plain `sanitize_sheet_title` lookup would have resolved to
        # before this map existed.
        name_map.setdefault(name, ws.title)

        row_idx = 0
        for table in worksheet.findall(_Q_TABLE):
            for row in table.findall(_Q_ROW):
                # SpreadsheetML omits empty rows/cells entirely and jumps with
                # a 1-based ss:Index, so appending sequentially would shift the
                # whole grid. Track cursors and write explicit coordinates.
                row_idx = int(row.get(_A_INDEX, row_idx + 1))

                col_idx = 0
                for cell in row.findall(_Q_CELL):
                    col_idx = int(cell.get(_A_INDEX, col_idx + 1))

                    value = _cell_value(cell)
                    if value is not None:
                        ws.cell(row=row_idx, column=col_idx, value=value)

                    # A merged cell spans MergeAcross + 1 columns; the ones it
                    # swallows must not be reused by the next cell.
                    col_idx += int(cell.get(_A_MERGE_ACROSS, 0))

        # Note: columns to the right of the last value-bearing cell are not
        # materialised — openpyxl drops style-less empty cells on save, so the
        # sheet ends at the widest populated column. Nothing is lost (those
        # columns are empty by definition) and interior gaps stay intact.

    if not wb.sheetnames:
        wb.create_sheet(title="Sheet1")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue(), name_map
