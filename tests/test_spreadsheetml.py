import io

import pytest
from openpyxl import load_workbook

from gex_msgraph._spreadsheetml import (
    looks_like_spreadsheetml,
    prepare_excel_bytes,
    to_xlsx_bytes,
)

_HEADER = (
    '<?xml version="1.0" encoding="UTF-16"?>\n'
    '<?mso-application progid="Excel.Sheet"?>\n'
    '<Workbook xmlns="urn:schemas-microsoft-com:office:spreadsheet"\n'
    ' xmlns:o="urn:schemas-microsoft-com:office:office"\n'
    ' xmlns:x="urn:schemas-microsoft-com:office:excel"\n'
    ' xmlns:ss="urn:schemas-microsoft-com:office:spreadsheet"\n'
    ' xmlns:html="http://www.w3.org/TR/REC-html40">\n'
)


def build(*sheets: str, encoding: str = "utf-16-be") -> bytes:
    """Wrap `<Worksheet>` fragments into a full document, BOM-encoded."""
    doc = _HEADER + "".join(sheets) + "</Workbook>"
    bom = {"utf-16-be": "﻿", "utf-16-le": "﻿", "utf-8-sig": ""}
    return (bom.get(encoding, "") + doc).encode(encoding)


def sheet(rows: str, name: str = "Sheet1") -> str:
    return f'<Worksheet ss:Name="{name}"><Table>{rows}</Table></Worksheet>'


def row(cells: str, index: int | None = None) -> str:
    attr = f' ss:Index="{index}"' if index is not None else ""
    return f"<Row{attr}>{cells}</Row>"


def cell(
    value: str,
    *,
    type_: str = "String",
    index: int | None = None,
    merge_across: int | None = None,
) -> str:
    attrs = ""
    if index is not None:
        attrs += f' ss:Index="{index}"'
    if merge_across is not None:
        attrs += f' ss:MergeAcross="{merge_across}"'
    return f'<Cell{attrs}><Data ss:Type="{type_}">{value}</Data></Cell>'


def grid(data: bytes, sheet_name: str | None = None) -> list[tuple]:
    """Convert to xlsx and read the resulting sheet back as rows of values."""
    wb = load_workbook(io.BytesIO(to_xlsx_bytes(data)))
    ws = wb[sheet_name] if sheet_name else wb[wb.sheetnames[0]]
    return list(ws.values)


def test_decodes_utf16be_with_bom():
    data = build(sheet(row(cell("hello"))))
    assert data.startswith(b"\xfe\xff")
    assert grid(data) == [("hello",)]


@pytest.mark.parametrize("encoding", ["utf-16-be", "utf-16-le", "utf-8-sig", "utf-8"])
def test_decodes_every_supported_encoding(encoding):
    assert grid(build(sheet(row(cell("hello"))), encoding=encoding)) == [("hello",)]


def test_unescaped_ampersand_is_repaired():
    # Real SAP output: bare "&" inside a String cell, which is not well-formed
    # XML and makes ElementTree raise without the repair pass.
    data = build(sheet(row(cell("NH TMCP ĐT&PT Việt Nam"))))
    assert grid(data) == [("NH TMCP ĐT&PT Việt Nam",)]


def test_existing_entities_are_not_double_escaped():
    data = build(sheet(row(cell("A&amp;B &lt;x&gt; &#65;"))))
    assert grid(data) == [("A&B <x> A",)]


def test_row_index_skips_rows():
    data = build(sheet(row(cell("top")) + row(cell("bottom"), index=4)))
    assert grid(data) == [
        ("top",),
        (None,),
        (None,),
        ("bottom",),
    ]


def test_cell_index_skips_columns():
    data = build(sheet(row(cell("a") + cell("d", index=4))))
    assert grid(data) == [("a", None, None, "d")]


def test_merge_across_consumes_columns():
    # A cell with MergeAcross="2" occupies 3 columns, so the next cell lands in
    # column 4 — not column 2.
    data = build(sheet(row(cell("wide", merge_across=2) + cell("after"))))
    assert grid(data) == [("wide", None, None, "after")]


def test_merge_across_combined_with_index():
    # Mirrors a real file: <Cell ss:Index="8" ss:MergeAcross="2"> spans 8→10.
    data = build(sheet(row(cell("m", index=8, merge_across=2) + cell("next"))))
    values = grid(data)[0]
    assert values[7] == "m"
    assert values[10] == "next"
    assert len(values) == 11


def test_number_type_becomes_float():
    data = build(sheet(row(cell("210491420093.00", type_="Number") + cell("x"))))
    assert grid(data) == [(210491420093.0, "x")]


def test_unparsable_number_falls_back_to_string():
    data = build(sheet(row(cell("n/a", type_="Number"))))
    assert grid(data) == [("n/a",)]


def test_empty_data_becomes_none():
    data = build(sheet(row('<Cell><Data ss:Type="String"></Data></Cell>' + cell("b"))))
    assert grid(data) == [(None, "b")]


def test_cell_without_data_element_becomes_none():
    data = build(sheet(row("<Cell/>" + cell("b"))))
    assert grid(data) == [(None, "b")]


def test_rich_text_runs_are_concatenated():
    rich = (
        '<Cell><Data ss:Type="String">'
        "plain<html:B xmlns:html=\"http://www.w3.org/TR/REC-html40\">bold</html:B>tail"
        "</Data></Cell>"
    )
    assert grid(build(sheet(row(rich)))) == [("plainboldtail",)]


def test_short_last_row_does_not_narrow_the_grid():
    data = build(sheet(row(cell("a") + cell("d", index=4)) + row(cell("x"))))
    assert grid(data) == [
        ("a", None, None, "d"),
        ("x", None, None, None),
    ]


def test_cdata_content_is_left_alone():
    # "&" inside CDATA is already valid XML; escaping it would corrupt the value.
    rich = '<Cell><Data ss:Type="String"><![CDATA[a & b]]></Data></Cell>'
    assert grid(build(sheet(row(rich)))) == [("a & b",)]


def test_bare_ampersand_outside_cdata_still_repaired_when_cdata_present():
    cells = (
        '<Cell><Data ss:Type="String"><![CDATA[x & y]]></Data></Cell>'
        '<Cell><Data ss:Type="String">p & q</Data></Cell>'
    )
    assert grid(build(sheet(row(cells)))) == [("x & y", "p & q")]


def test_entity_expansion_is_neutralised():
    # ElementTree does expand internal entities, so a billion-laughs document
    # would blow up memory. The ampersand repair turns every non-predefined
    # reference into literal text, which defuses it — lock that in.
    depth = 6
    entities = '<!ENTITY a0 "AAAAAAAAAA">' + "".join(
        f'<!ENTITY a{i} "{"&a%d;" % (i - 1) * 10}">' for i in range(1, depth + 1)
    )
    doc = (
        f"<!DOCTYPE Workbook [{entities}]>"
        + _HEADER.split("?>\n", 2)[-1]
        + sheet(row('<Cell><Data ss:Type="String">&a6;</Data></Cell>'))
        + "</Workbook>"
    )
    out = grid(("﻿" + doc).encode("utf-16-be"))
    assert out == [("&a6;",)]


def test_illegal_sheet_name_characters_are_sanitised():
    data = build(sheet(row(cell("v")), name="Q1/Q2:[x]"))
    wb = load_workbook(io.BytesIO(to_xlsx_bytes(data)))
    assert wb.sheetnames == ["Q1_Q2__x_"]
    assert list(wb["Q1_Q2__x_"].values) == [("v",)]


def test_overlong_sheet_name_is_truncated():
    data = build(sheet(row(cell("v")), name="S" * 40))
    assert load_workbook(io.BytesIO(to_xlsx_bytes(data))).sheetnames == ["S" * 31]


def test_undecodable_bytes_fall_back_with_a_warning(caplog):
    # One bad code unit inside a value must not cost the whole file, but it
    # must not pass silently either.
    data = build(sheet(row(cell("aQc"))))
    corrupt = data.replace(b"\x00Q", b"\xd8\x00", 1)  # unpaired high surrogate
    assert corrupt != data

    with caplog.at_level("WARNING", logger="gex_msgraph"):
        assert grid(corrupt) == [("a�c",)]
    assert "decode failed" in caplog.text


def test_multiple_worksheets_keep_their_names():
    data = build(
        sheet(row(cell("one")), name="First"),
        sheet(row(cell("two")), name="Second"),
    )
    wb = load_workbook(io.BytesIO(to_xlsx_bytes(data)))
    assert wb.sheetnames == ["First", "Second"]
    assert grid(data, "Second") == [("two",)]


def test_workbook_without_worksheets_still_produces_a_sheet():
    wb = load_workbook(io.BytesIO(to_xlsx_bytes(build())))
    assert wb.sheetnames == ["Sheet1"]


def test_pandas_can_read_the_converted_bytes():
    import pandas as pd

    data = build(
        sheet(
            row(cell("acct") + cell("amount"))
            + row(cell("111") + cell("1500.5", type_="Number"))
        )
    )
    df = pd.read_excel(io.BytesIO(to_xlsx_bytes(data)))
    assert list(df.columns) == ["acct", "amount"]
    assert df["amount"].tolist() == [1500.5]


def test_looks_like_spreadsheetml_true():
    assert looks_like_spreadsheetml(build(sheet(row(cell("a")))))


@pytest.mark.parametrize("encoding", ["utf-16-be", "utf-16-le", "utf-8-sig", "utf-8"])
def test_looks_like_spreadsheetml_true_for_every_encoding(encoding):
    assert looks_like_spreadsheetml(build(sheet(row(cell("a"))), encoding=encoding))


def test_looks_like_spreadsheetml_false_for_real_xlsx():
    from openpyxl import Workbook

    wb = Workbook()
    wb.active["A1"] = "hello"
    buf = io.BytesIO()
    wb.save(buf)
    assert buf.getvalue().startswith(b"PK\x03\x04")
    assert not looks_like_spreadsheetml(buf.getvalue())


def test_looks_like_spreadsheetml_false_for_ole2_xls():
    assert not looks_like_spreadsheetml(b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1" + b"\x00" * 64)


def test_looks_like_spreadsheetml_false_for_csv():
    assert not looks_like_spreadsheetml(b"a,b,c\n1,2,3\n")


def test_looks_like_spreadsheetml_false_for_unrelated_xml():
    assert not looks_like_spreadsheetml(b'<?xml version="1.0"?><root><a/></root>')


def test_looks_like_spreadsheetml_false_for_empty():
    assert not looks_like_spreadsheetml(b"")


def test_prepare_excel_bytes_passes_through_xlsx():
    from openpyxl import Workbook

    wb = Workbook()
    wb.active["A1"] = "hello"
    buf = io.BytesIO()
    wb.save(buf)
    data = buf.getvalue()
    out, selector = prepare_excel_bytes(data, "Sheet")
    assert out is data
    assert selector == "Sheet"


def test_prepare_excel_bytes_converts_and_remaps_sheet():
    data = build(sheet(row(cell("v")), name="Q1/Q2:[x]"))
    out, selector = prepare_excel_bytes(data, "Q1/Q2:[x]")
    assert not looks_like_spreadsheetml(out)
    assert selector == "Q1_Q2__x_"
    assert load_workbook(io.BytesIO(out)).sheetnames == ["Q1_Q2__x_"]


def test_prepare_excel_bytes_skips_sanitize_when_disabled():
    # Glob patterns go through untouched: "[" and "]" are pattern syntax.
    data = build(sheet(row(cell("v")), name="Q1/Q2:[x]"))
    _, selector = prepare_excel_bytes(data, "Q1/Q2:[x]", sanitize_sheet_name=False)
    assert selector == "Q1/Q2:[x]"


def test_prepare_excel_bytes_resolves_sanitize_collision_to_the_right_sheet():
    # Two distinct source names that sanitize to the same title: openpyxl
    # dedupes the second (Q1_A -> Q1_A1). A blind re-sanitize of the request
    # would resolve both to "Q1_A" and silently hand back sheet 1's data for
    # a request naming sheet 2 — the map must catch this.
    data = build(
        sheet(row(cell("one")), name="Q1/A"),
        sheet(row(cell("two")), name="Q1:A"),
    )
    _, first = prepare_excel_bytes(data, "Q1/A")
    _, second = prepare_excel_bytes(data, "Q1:A")
    assert first == "Q1_A"
    assert second == "Q1_A1"
    assert first != second

    out, _ = prepare_excel_bytes(data, "Q1:A")
    assert list(load_workbook(io.BytesIO(out))[second].values) == [("two",)]


def test_prepare_excel_bytes_duplicate_source_names_resolve_to_the_first():
    # A source that repeats a worksheet name is already malformed (Excel
    # refuses to save that), but must not crash — and the map must not let
    # the second occurrence silently steal the name from the first.
    data = build(
        sheet(row(cell("one")), name="Sheet1"),
        sheet(row(cell("two")), name="Sheet1"),
    )
    out, selector = prepare_excel_bytes(data, "Sheet1")
    assert selector == "Sheet1"
    assert list(load_workbook(io.BytesIO(out))[selector].values) == [("one",)]


def test_prepare_excel_bytes_unknown_name_falls_back_to_blind_sanitize():
    # A typo / name that never appeared in the file: fall back rather than
    # KeyError, so pandas still raises its normal "sheet not found".
    data = build(sheet(row(cell("v")), name="Sheet1"))
    _, selector = prepare_excel_bytes(data, "Nope")
    assert selector == "Nope"


def test_prepare_excel_bytes_leaves_positional_sheet_alone():
    data = build(sheet(row(cell("v")), name="Q1/Q2:[x]"))
    _, selector = prepare_excel_bytes(data, 0)
    assert selector == 0


def test_looks_like_spreadsheetml_ignores_namespace_beyond_sniff_window():
    # Namespace pushed past the 8 KiB window by a huge comment: we would rather
    # miss it and fall through to pandas than scan whole multi-MB downloads.
    padded = _HEADER.replace(
        "<Workbook", "<!--" + "x" * 9000 + "-->\n<Workbook"
    ) + "</Workbook>"
    assert not looks_like_spreadsheetml(("﻿" + padded).encode("utf-16-be"))
