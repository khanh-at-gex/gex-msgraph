import base64
import json
import pytest
import httpx
import respx
import asyncio
from gex_msgraph import GraphClient


async def test_init_missing_env(env_vars, monkeypatch):
    monkeypatch.delenv("MS_DAS_U1_CLIENT_ID")
    with pytest.raises(KeyError):
        GraphClient("das_u1")


async def test_init_username_without_password_raises(env_vars, monkeypatch):
    monkeypatch.delenv("MS_DAS_U1_PASSWORD")
    with pytest.raises(ValueError, match="both username and password"):
        GraphClient("das_u1")


async def test_init_app_only_when_username_and_password_absent(env_vars, monkeypatch):
    monkeypatch.delenv("MS_DAS_U1_USERNAME")
    monkeypatch.delenv("MS_DAS_U1_PASSWORD")
    monkeypatch.setenv("MS_DAS_U1_DEFAULT_DRIVE_ID", "b!abc")

    captured = {}

    class FakeApp:
        def acquire_token_for_client(self, scopes=None):
            captured["called"] = True
            return {"access_token": "app-only-token"}

    monkeypatch.setattr(
        "gex_msgraph._core.msal.ConfidentialClientApplication",
        lambda *a, **kw: FakeApp(),
    )
    client = GraphClient("das_u1")
    assert client._provider.get_token() == "app-only-token"
    assert captured["called"]


async def test_init_app_only_requires_default_drive_id(env_vars, monkeypatch):
    monkeypatch.delenv("MS_DAS_U1_USERNAME")
    monkeypatch.delenv("MS_DAS_U1_PASSWORD")
    with pytest.raises(ValueError, match="default_drive_id"):
        GraphClient("das_u1")


async def test_init_ropc_uses_username_password_flow(env_vars, monkeypatch):
    captured = {}

    class FakeApp:
        def get_accounts(self, username=None):
            return []

        def acquire_token_by_username_password(self, username, password, scopes):
            captured["username"] = username
            return {"access_token": "ropc-token"}

        def acquire_token_for_client(self, scopes=None):
            raise AssertionError("app-only flow must not be used when creds present")

    monkeypatch.setattr(
        "gex_msgraph._core.msal.ConfidentialClientApplication",
        lambda *a, **kw: FakeApp(),
    )
    client = GraphClient("das_u1")
    assert client._provider.get_token() == "ropc-token"
    assert captured["username"] == "fake_user"


async def test_init_rejects_default_site_id_kwarg(env_vars):
    # Removed in v0.3.0 — passing it must fail loudly, not be silently dropped.
    with pytest.raises(TypeError):
        GraphClient("das_u1", default_site_id="site123")


async def test_download(env_vars, mock_token):
    with respx.mock(assert_all_called=False) as rm:
        rm.get("https://graph.microsoft.com/v1.0/me/drive/root:/file.txt").mock(
            return_value=httpx.Response(
                200, json={"@microsoft.graph.downloadUrl": "https://download.url"}
            )
        )
        rm.get("https://download.url").mock(
            return_value=httpx.Response(200, content=b"hello bytes")
        )

        client = GraphClient("das_u1")
        async with client:
            res = await client.download(item_path="file.txt")
            assert res == b"hello bytes"


async def test_retry_429(env_vars, mock_token):
    with respx.mock(assert_all_called=False) as rm:
        route = rm.get("https://graph.microsoft.com/v1.0/me/drive/root:/file.txt")
        route.side_effect = [
            httpx.Response(429, headers={"Retry-After": "0"}),
            httpx.Response(
                200, json={"@microsoft.graph.downloadUrl": "https://download.url"}
            ),
        ]
        rm.get("https://download.url").mock(
            return_value=httpx.Response(200, content=b"ok")
        )

        client = GraphClient("das_u1")
        async with client:
            res = await client.download(item_path="file.txt")
            assert res == b"ok"
            assert route.call_count == 2


async def test_retry_503(env_vars, mock_token):
    with respx.mock(assert_all_called=False) as rm:
        route = rm.get("https://graph.microsoft.com/v1.0/me/drive/root:/file.txt")
        route.mock(return_value=httpx.Response(503))

        client = GraphClient("das_u1")

        async with client:
            with pytest.raises(httpx.HTTPStatusError):
                # Will retry 3 times plus initial = 4 calls.
                # However, backoff can take a bit if it uses real sleep.
                # We can mock sleep.
                async def mock_sleep(x):
                    pass

                with pytest.MonkeyPatch.context() as m:
                    m.setattr(asyncio, "sleep", mock_sleep)
                    await client.download(item_path="file.txt")

            assert route.call_count == 4


async def test_no_retry_403(env_vars, mock_token):
    with respx.mock(assert_all_called=False) as rm:
        route = rm.get("https://graph.microsoft.com/v1.0/me/drive/root:/file.txt")
        route.mock(return_value=httpx.Response(403))

        client = GraphClient("das_u1")
        async with client:
            with pytest.raises(httpx.HTTPStatusError):
                await client.download(item_path="file.txt")

            assert route.call_count == 1


async def test_download_retries_transient_error_on_download_url(env_vars, mock_token):
    with respx.mock(assert_all_called=False) as rm:
        rm.get("https://graph.microsoft.com/v1.0/me/drive/root:/file.txt").mock(
            return_value=httpx.Response(
                200, json={"@microsoft.graph.downloadUrl": "https://download.url"}
            )
        )
        route = rm.get("https://download.url")
        route.side_effect = [
            httpx.Response(503),
            httpx.Response(200, content=b"hello bytes"),
        ]

        client = GraphClient("das_u1")
        async with client:
            with pytest.MonkeyPatch.context() as m:
                async def mock_sleep(x):
                    pass
                m.setattr(asyncio, "sleep", mock_sleep)
                res = await client.download(item_path="file.txt")
            assert res == b"hello bytes"
            assert route.call_count == 2


async def test_download_no_retry_on_404_from_download_url(env_vars, mock_token):
    with respx.mock(assert_all_called=False) as rm:
        rm.get("https://graph.microsoft.com/v1.0/me/drive/root:/file.txt").mock(
            return_value=httpx.Response(
                200, json={"@microsoft.graph.downloadUrl": "https://download.url"}
            )
        )
        route = rm.get("https://download.url")
        route.mock(return_value=httpx.Response(404))

        client = GraphClient("das_u1")
        async with client:
            with pytest.raises(httpx.HTTPStatusError):
                await client.download(item_path="file.txt")
            assert route.call_count == 1


async def test_walk_pagination(env_vars, mock_token):
    with respx.mock(assert_all_called=False) as rm:
        route1 = rm.get("https://graph.microsoft.com/v1.0/me/drive/root/children")
        route1.mock(
            return_value=httpx.Response(
                200,
                json={
                    "value": [{"name": "file1.txt", "size": 10}],
                    "@odata.nextLink": "https://graph.microsoft.com/v1.0/page2",
                },
            )
        )

        route2 = rm.get("https://graph.microsoft.com/v1.0/page2")
        route2.mock(
            return_value=httpx.Response(
                200, json={"value": [{"name": "file2.txt", "size": 20}]}
            )
        )

        client = GraphClient("das_u1")
        async with client:
            files = await client.walk()
            assert len(files) == 2
            assert files[0].name == "file1.txt"
            assert files[1].name == "file2.txt"


async def test_send_mail(env_vars, mock_token):
    with respx.mock(assert_all_called=False) as rm:
        route = rm.post("https://graph.microsoft.com/v1.0/me/sendMail")
        route.mock(return_value=httpx.Response(202))

        client = GraphClient("das_u1")
        async with client:
            await client.send_mail(to="test@test.com", subject="Subj", body="Hello")

        req = route.calls[0].request
        payload = req.read().decode("utf-8")
        assert "test@test.com" in payload
        assert "Subj" in payload


async def test_send_mail_html(env_vars, mock_token):
    with respx.mock(assert_all_called=False) as rm:
        route = rm.post("https://graph.microsoft.com/v1.0/me/sendMail")
        route.mock(return_value=httpx.Response(202))
        async with GraphClient("das_u1") as client:
            await client.send_mail(
                to="a@b.com", subject="S", body="<b>Hi</b>", body_type="html"
            )
    payload = json.loads(route.calls[0].request.content)
    assert payload["message"]["body"]["contentType"] == "HTML"


async def test_send_mail_attachment_path(env_vars, mock_token, tmp_path):
    f = tmp_path / "report.txt"
    f.write_bytes(b"hello")
    with respx.mock(assert_all_called=False) as rm:
        route = rm.post("https://graph.microsoft.com/v1.0/me/sendMail")
        route.mock(return_value=httpx.Response(202))
        async with GraphClient("das_u1") as client:
            await client.send_mail(to="a@b.com", subject="S", body="body", attachments=[f])
    payload = json.loads(route.calls[0].request.content)
    att = payload["message"]["attachments"][0]
    assert att["name"] == "report.txt"
    assert base64.b64decode(att["contentBytes"]) == b"hello"


async def test_send_mail_attachment_bytes(env_vars, mock_token):
    with respx.mock(assert_all_called=False) as rm:
        route = rm.post("https://graph.microsoft.com/v1.0/me/sendMail")
        route.mock(return_value=httpx.Response(202))
        async with GraphClient("das_u1") as client:
            await client.send_mail(
                to="a@b.com", subject="S", body="body",
                attachments=[("data.csv", b"col1,col2\n1,2")],
            )
    payload = json.loads(route.calls[0].request.content)
    att = payload["message"]["attachments"][0]
    assert att["name"] == "data.csv"
    assert att["contentType"]  # platform-dependent; just ensure it is set
    assert base64.b64decode(att["contentBytes"]) == b"col1,col2\n1,2"


async def test_send_mail_attachment_sharepoint(env_vars, mock_token):
    with respx.mock(assert_all_called=False) as rm:
        rm.get("https://graph.microsoft.com/v1.0/me/drive/root:/Reports/Q1.xlsx").mock(
            return_value=httpx.Response(
                200, json={"@microsoft.graph.downloadUrl": "https://cdn.example.com/Q1.xlsx"}
            )
        )
        rm.get("https://cdn.example.com/Q1.xlsx").mock(
            return_value=httpx.Response(200, content=b"xlsxbytes")
        )
        route = rm.post("https://graph.microsoft.com/v1.0/me/sendMail")
        route.mock(return_value=httpx.Response(202))
        async with GraphClient("das_u1") as client:
            await client.send_mail(
                to="a@b.com", subject="S", body="body",
                attachments=[{"item_path": "Reports/Q1.xlsx"}],
            )
    payload = json.loads(route.calls[0].request.content)
    att = payload["message"]["attachments"][0]
    assert att["name"] == "Q1.xlsx"
    assert base64.b64decode(att["contentBytes"]) == b"xlsxbytes"


async def test_send_teams_message(env_vars, mock_token):
    with respx.mock(assert_all_called=False) as rm:
        route = rm.post(
            "https://graph.microsoft.com/v1.0/teams/t1/channels/c1/messages"
        )
        route.mock(return_value=httpx.Response(201))

        client = GraphClient("das_u1")
        async with client:
            await client.send_teams_message(
                team_id="t1", channel_id="c1", text="Hello Teams"
            )

        req = route.calls[0].request
        payload = req.read().decode("utf-8")
        assert "Hello Teams" in payload


async def test_copy_file(env_vars, mock_token):
    with respx.mock(assert_all_called=False) as rm:
        route = rm.post("https://graph.microsoft.com/v1.0/me/drive/root:/src.xlsx:/copy")
        route.mock(return_value=httpx.Response(202))
        async with GraphClient("das_u1") as client:
            result = await client.copy_file(
                item_path="src.xlsx",
                dest_folder_path="Archive",
                new_name="src_copy.xlsx",
            )
    assert result is None  # wait=False keeps the fire-and-forget contract
    payload = json.loads(route.calls[0].request.content)
    assert "Archive" in payload["parentReference"]["path"]
    assert payload["name"] == "src_copy.xlsx"


async def test_copy_file_by_item_id(env_vars, mock_token):
    with respx.mock(assert_all_called=False) as rm:
        route = rm.post("https://graph.microsoft.com/v1.0/me/drive/items/01ABC/copy")
        route.mock(return_value=httpx.Response(202))
        async with GraphClient("das_u1") as client:
            await client.copy_file(item_id="01ABC", dest_folder_path="Archive")
    assert route.call_count == 1


async def test_copy_file_wait_polls_until_completed(env_vars, mock_token):
    with respx.mock(assert_all_called=False) as rm:
        rm.post("https://graph.microsoft.com/v1.0/me/drive/root:/src.xlsx:/copy").mock(
            return_value=httpx.Response(
                202, headers={"Location": "https://monitor.url/job1"}
            )
        )
        monitor = rm.get("https://monitor.url/job1")
        monitor.side_effect = [
            httpx.Response(202, json={"status": "inProgress"}),
            httpx.Response(
                200, json={"status": "completed", "resourceId": "NEW123"}
            ),
        ]
        rm.get("https://graph.microsoft.com/v1.0/me/drive/items/NEW123").mock(
            return_value=httpx.Response(
                200, json={"id": "NEW123", "name": "src_copy.xlsx", "size": 9}
            )
        )

        async with GraphClient("das_u1") as client:
            with pytest.MonkeyPatch.context() as m:
                async def mock_sleep(x):
                    pass
                m.setattr(asyncio, "sleep", mock_sleep)
                item = await client.copy_file(
                    item_path="src.xlsx", dest_folder_path="Archive", wait=True
                )

    assert item is not None and item.id == "NEW123"
    assert monitor.call_count == 2


async def test_copy_file_wait_survives_transient_monitor_error(env_vars, mock_token):
    with respx.mock(assert_all_called=False) as rm:
        rm.post("https://graph.microsoft.com/v1.0/me/drive/root:/src.xlsx:/copy").mock(
            return_value=httpx.Response(
                202, headers={"Location": "https://monitor.url/job3"}
            )
        )
        monitor = rm.get("https://monitor.url/job3")
        monitor.side_effect = [
            httpx.Response(503),  # transient — must keep polling, not TimeoutError
            httpx.Response(200, json={"status": "completed", "resourceId": "NEW9"}),
        ]
        rm.get("https://graph.microsoft.com/v1.0/me/drive/items/NEW9").mock(
            return_value=httpx.Response(200, json={"id": "NEW9", "name": "c.xlsx"})
        )

        async with GraphClient("das_u1") as client:
            with pytest.MonkeyPatch.context() as m:
                async def mock_sleep(x):
                    pass
                m.setattr(asyncio, "sleep", mock_sleep)
                item = await client.copy_file(
                    item_path="src.xlsx", dest_folder_path="Archive", wait=True
                )
    assert item is not None and item.id == "NEW9"
    assert monitor.call_count == 2


async def test_copy_file_wait_raises_on_monitor_4xx(env_vars, mock_token):
    from gex_msgraph import NotFoundError

    with respx.mock(assert_all_called=False) as rm:
        rm.post("https://graph.microsoft.com/v1.0/me/drive/root:/src.xlsx:/copy").mock(
            return_value=httpx.Response(
                202, headers={"Location": "https://monitor.url/job4"}
            )
        )
        rm.get("https://monitor.url/job4").mock(return_value=httpx.Response(404))

        async with GraphClient("das_u1") as client:
            with pytest.raises(NotFoundError):
                await client.copy_file(
                    item_path="src.xlsx", dest_folder_path="Archive", wait=True
                )


async def test_copy_file_wait_raises_on_failed_status(env_vars, mock_token):
    from gex_msgraph import GraphError

    with respx.mock(assert_all_called=False) as rm:
        rm.post("https://graph.microsoft.com/v1.0/me/drive/root:/src.xlsx:/copy").mock(
            return_value=httpx.Response(
                202, headers={"Location": "https://monitor.url/job2"}
            )
        )
        rm.get("https://monitor.url/job2").mock(
            return_value=httpx.Response(
                200,
                json={"status": "failed", "error": {"message": "quota exceeded"}},
            )
        )
        async with GraphClient("das_u1") as client:
            with pytest.raises(GraphError, match="quota exceeded"):
                await client.copy_file(
                    item_path="src.xlsx", dest_folder_path="Archive", wait=True
                )


async def test_exists_true(env_vars, mock_token):
    with respx.mock(assert_all_called=False) as rm:
        rm.get("https://graph.microsoft.com/v1.0/me/drive/root:/file.txt").mock(
            return_value=httpx.Response(200, json={"name": "file.txt", "size": 1})
        )
        async with GraphClient("das_u1") as client:
            assert await client.exists(item_path="file.txt") is True


async def test_exists_false(env_vars, mock_token):
    with respx.mock(assert_all_called=False) as rm:
        rm.get("https://graph.microsoft.com/v1.0/me/drive/root:/missing.txt").mock(
            return_value=httpx.Response(404)
        )
        async with GraphClient("das_u1") as client:
            assert await client.exists(item_path="missing.txt") is False


async def test_upload_many(env_vars, mock_token, tmp_path):
    f1 = tmp_path / "a.txt"
    f2 = tmp_path / "b.txt"
    f1.write_bytes(b"aaa")
    f2.write_bytes(b"bbb")
    with respx.mock(assert_all_called=False) as rm:
        rm.put("https://graph.microsoft.com/v1.0/me/drive/root:/remote/a.txt:/content").mock(
            return_value=httpx.Response(201, json={"name": "a.txt"})
        )
        rm.put("https://graph.microsoft.com/v1.0/me/drive/root:/remote/b.txt:/content").mock(
            return_value=httpx.Response(201, json={"name": "b.txt"})
        )
        async with GraphClient("das_u1") as client:
            results = await client.upload_many(
                [(f1, "remote/a.txt"), (f2, "remote/b.txt")]
            )
    assert len(results) == 2
    assert results[0]["name"] == "a.txt"


async def test_upload_large_file_uses_session(env_vars, mock_token, tmp_path, monkeypatch):
    # Shrink thresholds so a tiny file exercises the chunked path.
    monkeypatch.setattr("gex_msgraph._core._UPLOAD_SESSION_THRESHOLD", 10)
    monkeypatch.setattr("gex_msgraph._core._UPLOAD_CHUNK_SIZE", 64)

    big = tmp_path / "big.bin"
    big.write_bytes(b"x" * 100)  # 100 bytes -> 2 chunks of 64 + 36

    with respx.mock(assert_all_called=False) as rm:
        rm.post(
            "https://graph.microsoft.com/v1.0/me/drive/root:/remote/big.bin:/createUploadSession"
        ).mock(
            return_value=httpx.Response(200, json={"uploadUrl": "https://upload.url/s1"})
        )
        chunk_route = rm.put("https://upload.url/s1")
        chunk_route.side_effect = [
            httpx.Response(202, json={"nextExpectedRanges": ["64-99"]}),
            httpx.Response(201, json={"id": "BIG1", "name": "big.bin", "size": 100}),
        ]

        async with GraphClient("das_u1") as client:
            result = await client.upload(big, "remote/big.bin")

    assert result["id"] == "BIG1"
    assert chunk_route.call_count == 2
    first, second = chunk_route.calls
    assert first.request.headers["Content-Range"] == "bytes 0-63/100"
    assert second.request.headers["Content-Range"] == "bytes 64-99/100"
    # Pre-authenticated session URL must NOT get a Bearer header.
    assert "Authorization" not in first.request.headers


async def test_upload_small_file_still_uses_simple_put(env_vars, mock_token, tmp_path):
    small = tmp_path / "small.txt"
    small.write_bytes(b"tiny")
    with respx.mock(assert_all_called=False) as rm:
        route = rm.put(
            "https://graph.microsoft.com/v1.0/me/drive/root:/remote/small.txt:/content"
        )
        route.mock(return_value=httpx.Response(201, json={"name": "small.txt"}))
        async with GraphClient("das_u1") as client:
            result = await client.upload(small, "remote/small.txt")
    assert result["name"] == "small.txt"
    assert route.call_count == 1


async def test_download_to_path_streams_to_disk(env_vars, mock_token, tmp_path):
    from pathlib import Path

    with respx.mock(assert_all_called=False) as rm:
        rm.get("https://graph.microsoft.com/v1.0/me/drive/root:/file.bin").mock(
            return_value=httpx.Response(
                200, json={"@microsoft.graph.downloadUrl": "https://download.url"}
            )
        )
        rm.get("https://download.url").mock(
            return_value=httpx.Response(200, content=b"streamed content")
        )

        dest = tmp_path / "out.bin"
        async with GraphClient("das_u1") as client:
            returned = await client.download(item_path="file.bin", to_path=dest)

    assert returned == Path(dest)
    assert dest.read_bytes() == b"streamed content"


async def test_download_to_path_retries_transient_error(env_vars, mock_token, tmp_path):
    with respx.mock(assert_all_called=False) as rm:
        rm.get("https://graph.microsoft.com/v1.0/me/drive/root:/file.bin").mock(
            return_value=httpx.Response(
                200, json={"@microsoft.graph.downloadUrl": "https://download.url"}
            )
        )
        route = rm.get("https://download.url")
        route.side_effect = [
            httpx.Response(503),
            httpx.Response(200, content=b"ok after retry"),
        ]

        dest = tmp_path / "out.bin"
        async with GraphClient("das_u1") as client:
            with pytest.MonkeyPatch.context() as m:
                async def mock_sleep(x):
                    pass
                m.setattr(asyncio, "sleep", mock_sleep)
                await client.download(item_path="file.bin", to_path=dest)

    assert dest.read_bytes() == b"ok after retry"
    assert route.call_count == 2


async def test_upload_many_respects_max_concurrent(env_vars, mock_token):
    concurrent = 0
    peak = 0

    async def fake_upload(local, remote):
        nonlocal concurrent, peak
        concurrent += 1
        peak = max(peak, concurrent)
        await asyncio.sleep(0.01)
        concurrent -= 1
        return {"name": str(remote)}

    client = GraphClient("das_u1")
    client.upload = fake_upload  # type: ignore[method-assign]
    items = [(f"local{i}.txt", f"remote{i}.txt") for i in range(4)]
    results = await client.upload_many(items, max_concurrent=2)

    assert len(results) == 4
    assert peak <= 2


async def test_get_share_link(env_vars, mock_token):
    with respx.mock(assert_all_called=False) as rm:
        route = rm.post(
            "https://graph.microsoft.com/v1.0/me/drive/root:/report.xlsx:/createLink"
        )
        route.mock(
            return_value=httpx.Response(
                200, json={"link": {"webUrl": "https://tenant.sharepoint.com/shared/link"}}
            )
        )
        async with GraphClient("das_u1") as client:
            url = await client.get_share_link(item_path="report.xlsx")
    assert url == "https://tenant.sharepoint.com/shared/link"


async def test_read_parquet(env_vars, mock_token):
    import io
    import pandas as pd

    df_orig = pd.DataFrame({"a": [1, 2], "b": [3, 4]})
    buf = io.BytesIO()
    df_orig.to_parquet(buf, index=False)
    parquet_bytes = buf.getvalue()

    with respx.mock(assert_all_called=False) as rm:
        rm.get("https://graph.microsoft.com/v1.0/me/drive/root:/data.parquet").mock(
            return_value=httpx.Response(
                200, json={"@microsoft.graph.downloadUrl": "https://dl.url/data.parquet"}
            )
        )
        rm.get("https://dl.url/data.parquet").mock(
            return_value=httpx.Response(200, content=parquet_bytes)
        )
        async with GraphClient("das_u1") as client:
            df = await client.read_parquet(item_path="data.parquet")
    assert list(df.columns) == ["a", "b"]
    assert len(df) == 2


async def test_list_mail(env_vars, mock_token):
    with respx.mock(assert_all_called=False) as rm:
        rm.get(
            "https://graph.microsoft.com/v1.0/me/mailFolders/inbox/messages"
            "?$top=5&$select=id,subject,from,receivedDateTime,bodyPreview,hasAttachments"
        ).mock(
            return_value=httpx.Response(
                200, json={"value": [{"id": "m1", "subject": "Hello"}]}
            )
        )
        async with GraphClient("das_u1") as client:
            msgs = await client.list_mail(limit=5)
    assert len(msgs) == 1
    assert msgs[0]["subject"] == "Hello"


async def test_search_files(env_vars, mock_token):
    with respx.mock(assert_all_called=False) as rm:
        rm.get(
            "https://graph.microsoft.com/v1.0/me/drive/root/search(q='budget')?$top=25"
        ).mock(
            return_value=httpx.Response(
                200, json={"value": [{"name": "budget.xlsx", "size": 100}]}
            )
        )
        async with GraphClient("das_u1") as client:
            results = await client.search_files("budget")
    assert len(results) == 1
    assert results[0].name == "budget.xlsx"


async def test_search_files_respects_limit_across_pages(env_vars, mock_token):
    with respx.mock(assert_all_called=False) as rm:
        page1_items = [{"name": f"file{i}.xlsx", "size": 1} for i in range(20)]
        page2_items = [{"name": f"file{i}.xlsx", "size": 1} for i in range(20, 40)]

        rm.get(
            "https://graph.microsoft.com/v1.0/me/drive/root/search(q='x')?$top=25"
        ).mock(
            return_value=httpx.Response(
                200,
                json={
                    "value": page1_items,
                    "@odata.nextLink": "https://graph.microsoft.com/v1.0/page2",
                },
            )
        )
        route2 = rm.get("https://graph.microsoft.com/v1.0/page2")
        route2.mock(return_value=httpx.Response(200, json={"value": page2_items}))

        async with GraphClient("das_u1") as client:
            results = await client.search_files("x", limit=25)

    assert len(results) == 25
    # Only enough of page 2 was consumed to reach the limit; the generator
    # stopped once 25 results were collected, but the page fetch itself
    # (which returns the whole page at once) still only happened once.
    assert route2.call_count == 1


def test_sync_wrappers(env_vars, mock_token):
    with respx.mock(assert_all_called=False) as rm:
        rm.get("https://graph.microsoft.com/v1.0/me/drive/root:/file.txt").mock(
            return_value=httpx.Response(
                200, json={"@microsoft.graph.downloadUrl": "https://download.url"}
            )
        )
        rm.get("https://download.url").mock(
            return_value=httpx.Response(200, content=b"hello bytes")
        )

        client = GraphClient("das_u1")
        try:
            res = client.download_sync(item_path="file.txt")
            assert res == b"hello bytes"
        finally:
            client.close_sync()


def test_sync_wrapper_called_twice_reuses_loop(env_vars, mock_token):
    # Regression: second asyncio.run() used to hit "attached to a different
    # event loop"; the persistent background loop makes repeat calls safe.
    with respx.mock(assert_all_called=False) as rm:
        rm.get("https://graph.microsoft.com/v1.0/me/drive/root:/file.txt").mock(
            return_value=httpx.Response(
                200, json={"@microsoft.graph.downloadUrl": "https://download.url"}
            )
        )
        rm.get("https://download.url").mock(
            return_value=httpx.Response(200, content=b"hello bytes")
        )

        client = GraphClient("das_u1")
        try:
            assert client.download_sync(item_path="file.txt") == b"hello bytes"
            assert client.download_sync(item_path="file.txt") == b"hello bytes"
        finally:
            client.close_sync()


async def test_async_close_after_sync_use_delegates_to_sync_loop(env_vars, mock_token):
    # Regression: `await client.close()` after *_sync usage used to aclose the
    # httpx client from the wrong event loop.
    def _use_sync(client):
        with respx.mock(assert_all_called=False) as rm:
            rm.get("https://graph.microsoft.com/v1.0/me/drive/root:/f.txt").mock(
                return_value=httpx.Response(
                    200, json={"@microsoft.graph.downloadUrl": "https://download.url"}
                )
            )
            rm.get("https://download.url").mock(
                return_value=httpx.Response(200, content=b"x")
            )
            assert client.download_sync(item_path="f.txt") == b"x"

    client = GraphClient("das_u1")
    await asyncio.to_thread(_use_sync, client)
    assert client._sync_loop is not None
    await client.close()  # must not raise; must tear down the background loop
    assert client._sync_loop is None


async def test_sync_wrapper_raises_inside_running_loop(env_vars, mock_token):
    from gex_msgraph import GraphSyncInLoopError

    client = GraphClient("das_u1")
    async with client:
        with pytest.raises(GraphSyncInLoopError, match="await client"):
            client.download_sync(item_path="file.txt")


async def test_get_metadata(env_vars, mock_token):
    with respx.mock(assert_all_called=False) as rm:
        route = rm.get("https://graph.microsoft.com/v1.0/me/drive/root:/file.txt")
        route.mock(
            return_value=httpx.Response(200, json={"name": "file.txt", "size": 123})
        )

        client = GraphClient("das_u1")
        async with client:
            meta = await client.get_metadata(item_path="file.txt")
            assert meta.name == "file.txt"
            assert meta.size == 123


async def test_delete_file(env_vars, mock_token):
    with respx.mock(assert_all_called=False) as rm:
        route = rm.delete("https://graph.microsoft.com/v1.0/me/drive/root:/file.txt")
        route.mock(return_value=httpx.Response(204))

        client = GraphClient("das_u1")
        async with client:
            await client.delete_file(item_path="file.txt")
            assert route.called


async def test_move_file(env_vars, mock_token):
    with respx.mock(assert_all_called=False) as rm:
        route = rm.patch("https://graph.microsoft.com/v1.0/me/drive/root:/file.txt")
        route.mock(return_value=httpx.Response(200, json={"name": "new.txt"}))

        client = GraphClient("das_u1")
        async with client:
            res = await client.move_file(
                item_path="file.txt",
                dest_folder_path="dest/folder",
                new_name="new.txt",
            )
            assert res.name == "new.txt"
            payload = route.calls[0].request.read().decode()
            assert "dest/folder" in payload
            assert "new.txt" in payload


async def test_move_file_by_item_id(env_vars, mock_token):
    with respx.mock(assert_all_called=False) as rm:
        route = rm.patch("https://graph.microsoft.com/v1.0/me/drive/items/01ABC")
        route.mock(return_value=httpx.Response(200, json={"name": "file.txt"}))

        async with GraphClient("das_u1") as client:
            res = await client.move_file(item_id="01ABC", dest_folder_path="Archive")
    assert res.name == "file.txt"
    payload = json.loads(route.calls[0].request.content)
    assert "Archive" in payload["parentReference"]["path"]


async def test_move_file_requires_dest_or_name(env_vars, mock_token):
    async with GraphClient("das_u1") as client:
        with pytest.raises(ValueError, match="dest_folder_path or new_name"):
            await client.move_file(item_path="file.txt")


async def test_move_file_requires_exactly_one_identifier(env_vars, mock_token):
    async with GraphClient("das_u1") as client:
        with pytest.raises(ValueError, match="exactly one"):
            await client.move_file(
                item_path="a.txt", item_id="01ABC", dest_folder_path="x"
            )


async def test_create_folder(env_vars, mock_token):
    with respx.mock(assert_all_called=False) as rm:
        route = rm.post(
            "https://graph.microsoft.com/v1.0/me/drive/root:/parent:/children"
        )
        route.mock(
            return_value=httpx.Response(201, json={"name": "child", "folder": {}})
        )

        client = GraphClient("das_u1")
        async with client:
            res = await client.create_folder("parent/child")
            assert res.name == "child"
            assert res.is_folder


async def test_list_excel_sheets(env_vars, mock_token):
    with respx.mock(assert_all_called=False) as rm:
        route = rm.get(
            "https://graph.microsoft.com/v1.0/me/drive/root:/file.xlsx:/workbook/worksheets"
        )
        route.mock(
            return_value=httpx.Response(
                200, json={"value": [{"name": "Sheet1"}, {"name": "Sheet2"}]}
            )
        )

        client = GraphClient("das_u1")
        async with client:
            res = await client.list_excel_sheets(item_path="file.xlsx")
            assert res == ["Sheet1", "Sheet2"]


async def test_get_folder_tree(env_vars, mock_token):
    with respx.mock(assert_all_called=False) as rm:
        rm.get("https://graph.microsoft.com/v1.0/me/drive/root:/root").mock(
            return_value=httpx.Response(200, json={"name": "root", "folder": {}})
        )
        rm.get("https://graph.microsoft.com/v1.0/me/drive/root:/root:/children").mock(
            return_value=httpx.Response(
                200,
                json={"value": [{"name": "file.txt"}, {"name": "sub", "folder": {}}]},
            )
        )
        rm.get(
            "https://graph.microsoft.com/v1.0/me/drive/root:/root/sub:/children"
        ).mock(return_value=httpx.Response(200, json={"value": []}))

        client = GraphClient("das_u1")
        async with client:
            tree = await client.get_folder_tree("root")
            assert tree.item is not None and tree.item.name == "root"
            assert len(tree.children) == 2
            assert tree.children[0].item.name == "file.txt"
            assert tree.children[1].item.name == "sub"


async def test_get_folder_tree_preserves_order_with_parallel_sibling_recursion(
    env_vars, mock_token
):
    with respx.mock(assert_all_called=False) as rm:
        rm.get("https://graph.microsoft.com/v1.0/me/drive/root:/root").mock(
            return_value=httpx.Response(200, json={"name": "root", "folder": {}})
        )
        rm.get("https://graph.microsoft.com/v1.0/me/drive/root:/root:/children").mock(
            return_value=httpx.Response(
                200,
                json={
                    "value": [
                        {"name": "subA", "folder": {}},
                        {"name": "subB", "folder": {}},
                    ]
                },
            )
        )
        rm.get(
            "https://graph.microsoft.com/v1.0/me/drive/root:/root/subA:/children"
        ).mock(
            return_value=httpx.Response(
                200, json={"value": [{"name": "a1.txt"}, {"name": "a2.txt"}]}
            )
        )
        rm.get(
            "https://graph.microsoft.com/v1.0/me/drive/root:/root/subB:/children"
        ).mock(return_value=httpx.Response(200, json={"value": [{"name": "b1.txt"}]}))

        client = GraphClient("das_u1")
        async with client:
            tree = await client.get_folder_tree("root")

        assert [c.item.name for c in tree.children] == ["subA", "subB"]
        sub_a, sub_b = tree.children
        assert [c.item.name for c in sub_a.children] == ["a1.txt", "a2.txt"]
        assert [c.item.name for c in sub_b.children] == ["b1.txt"]


async def test_get_teams_messages(env_vars, mock_token):
    with respx.mock(assert_all_called=False) as rm:
        route = rm.get(
            "https://graph.microsoft.com/v1.0/teams/t1/channels/c1/messages?$top=10"
        )
        route.mock(return_value=httpx.Response(200, json={"value": [{"id": "m1"}]}))

        client = GraphClient("das_u1")
        async with client:
            res = await client.get_teams_messages("t1", "c1")
            assert len(res) == 1
            assert res[0]["id"] == "m1"


async def test_explicit_drive_id_url_format(env_vars_with_drive, mock_token):
    """Bug 1 regression: explicit DEFAULT_DRIVE_ID emits /drives/{id}/..., not /me/drive/..."""
    with respx.mock(assert_all_called=False) as rm:
        route = rm.get("https://graph.microsoft.com/v1.0/drives/b!abc/root:/file.txt")
        route.mock(
            return_value=httpx.Response(200, json={"name": "file.txt", "size": 1})
        )

        client = GraphClient("das_u1")
        async with client:
            meta = await client.get_metadata(item_path="file.txt")
            assert meta.name == "file.txt"
            assert route.called


async def test_read_excel_many_missing_sheet_raise_overrides_on_error(
    env_vars, mock_token
):
    """Bug 4 regression: on_missing_sheet='raise' must raise even when on_error='skip'."""
    import io
    import openpyxl

    # Build a real one-sheet xlsx in memory so pd.ExcelFile can open it.
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "OnlySheet"
    ws["A1"] = "hello"
    buf = io.BytesIO()
    wb.save(buf)
    xlsx_bytes = buf.getvalue()

    with respx.mock(assert_all_called=False) as rm:
        rm.get("https://graph.microsoft.com/v1.0/me/drive/root:/f.xlsx").mock(
            return_value=httpx.Response(
                200, json={"@microsoft.graph.downloadUrl": "https://dl.url"}
            )
        )
        rm.get("https://dl.url").mock(
            return_value=httpx.Response(200, content=xlsx_bytes)
        )

        client = GraphClient("das_u1")
        async with client:
            with pytest.raises(ValueError, match="not found"):
                await client.read_excel_many(
                    ["f.xlsx"],
                    sheet="DoesNotExist",
                    on_missing_sheet="raise",
                    on_error="skip",
                )


async def test_list_excel_sheets_by_item_id(env_vars, mock_token):
    """Bug 3 regression: id-addressed list_excel_sheets uses /items/{id}/workbook/worksheets (no colon)."""
    with respx.mock(assert_all_called=False) as rm:
        route = rm.get(
            "https://graph.microsoft.com/v1.0/me/drive/items/xyz/workbook/worksheets"
        )
        route.mock(return_value=httpx.Response(200, json={"value": [{"name": "S1"}]}))

        client = GraphClient("das_u1")
        async with client:
            res = await client.list_excel_sheets(item_id="xyz")
            assert res == ["S1"]


async def test_list_excel_sheets_by_share_url(env_vars, mock_token):
    """Bug 3 regression: share-addressed list_excel_sheets uses /shares/{enc}/driveItem/workbook/worksheets."""
    from gex_msgraph._files import encode_share_url

    share = "https://tenant.sharepoint.com/sites/x/Shared%20Documents/f.xlsx"
    encoded = encode_share_url(share)

    with respx.mock(assert_all_called=False) as rm:
        route = rm.get(
            f"https://graph.microsoft.com/v1.0/shares/{encoded}/driveItem/workbook/worksheets"
        )
        route.mock(return_value=httpx.Response(200, json={"value": [{"name": "S1"}]}))

        client = GraphClient("das_u1")
        async with client:
            res = await client.list_excel_sheets(share_url=share)
            assert res == ["S1"]


async def test_read_excel_many_return_status(env_vars, mock_token):
    import io
    import openpyxl
    import pandas as pd

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "hello"
    ws["A2"] = "world"
    buf = io.BytesIO()
    wb.save(buf)
    xlsx_bytes = buf.getvalue()

    with respx.mock(assert_all_called=False) as rm:
        # File 1: Success
        rm.get("https://graph.microsoft.com/v1.0/me/drive/root:/f1.xlsx").mock(
            return_value=httpx.Response(
                200, json={"@microsoft.graph.downloadUrl": "https://dl.url/1"}
            )
        )
        rm.get("https://dl.url/1").mock(
            return_value=httpx.Response(200, content=xlsx_bytes)
        )

        # File 2: Error (simulating an HTTP 404 or bad resolve)
        rm.get("https://graph.microsoft.com/v1.0/me/drive/root:/f2.xlsx").mock(
            return_value=httpx.Response(404)
        )

        client = GraphClient("das_u1")
        async with client:
            df, status_df = await client.read_excel_many(
                ["f1.xlsx", "f2.xlsx"],
                sheet="Sheet1",
                on_error="skip",
                return_status=True,
            )

            assert isinstance(df, pd.DataFrame)
            assert isinstance(status_df, pd.DataFrame)
            assert len(df) == 1

            status_dict = status_df.set_index("path")["status"].to_dict()
            assert status_dict["f1.xlsx"] == "success"
            assert status_dict["f2.xlsx"] == "error"


async def test_read_csv_many_return_status(env_vars, mock_token):
    import pandas as pd

    csv_bytes = b"col1,col2\n1,2\n3,4"

    with respx.mock(assert_all_called=False) as rm:
        # File 1: Success
        rm.get("https://graph.microsoft.com/v1.0/me/drive/root:/f1.csv").mock(
            return_value=httpx.Response(
                200, json={"@microsoft.graph.downloadUrl": "https://dl.url/1"}
            )
        )
        rm.get("https://dl.url/1").mock(
            return_value=httpx.Response(200, content=csv_bytes)
        )

        # File 2: Error
        rm.get("https://graph.microsoft.com/v1.0/me/drive/root:/f2.csv").mock(
            return_value=httpx.Response(404)
        )

        client = GraphClient("das_u1")
        async with client:
            df, status_df = await client.read_csv_many(
                ["f1.csv", "f2.csv"], on_error="skip", return_status=True
            )

            assert isinstance(df, pd.DataFrame)
            assert len(df) == 2
            assert df["col1"].iloc[0] == 1

            status_dict = status_df.set_index("path")["status"].to_dict()
            assert status_dict["f1.csv"] == "success"
            assert status_dict["f2.csv"] == "error"


def _spreadsheetml_bytes() -> bytes:
    """A SAP-style ``.xls``: SpreadsheetML 2003, UTF-16BE, with a bare ``&``,
    a skipped column via ``ss:Index`` and a ``MergeAcross`` span."""
    doc = (
        '<?xml version="1.0" encoding="UTF-16"?>\n'
        '<?mso-application progid="Excel.Sheet"?>\n'
        '<Workbook xmlns="urn:schemas-microsoft-com:office:spreadsheet"'
        ' xmlns:ss="urn:schemas-microsoft-com:office:spreadsheet">'
        '<Worksheet ss:Name="Sheet1"><Table>'
        "<Row>"
        '<Cell><Data ss:Type="String">acct</Data></Cell>'
        '<Cell ss:Index="3"><Data ss:Type="String">name</Data></Cell>'
        '<Cell ss:MergeAcross="1"><Data ss:Type="String">amount</Data></Cell>'
        "</Row>"
        "<Row>"
        '<Cell><Data ss:Type="String">111</Data></Cell>'
        '<Cell ss:Index="3"><Data ss:Type="String">NH TMCP ĐT&PT Việt Nam</Data></Cell>'
        '<Cell ss:MergeAcross="1"><Data ss:Type="Number">210491420093.00</Data></Cell>'
        "</Row>"
        "</Table></Worksheet></Workbook>"
    )
    return ("﻿" + doc).encode("utf-16-be")


async def test_read_excel_handles_spreadsheetml(env_vars, mock_token):
    with respx.mock(assert_all_called=False) as rm:
        rm.get("https://graph.microsoft.com/v1.0/me/drive/root:/TB/jan.xls").mock(
            return_value=httpx.Response(
                200, json={"@microsoft.graph.downloadUrl": "https://dl.url/jan.xls"}
            )
        )
        rm.get("https://dl.url/jan.xls").mock(
            return_value=httpx.Response(200, content=_spreadsheetml_bytes())
        )
        async with GraphClient("das_u1") as client:
            df = await client.read_excel(item_path="TB/jan.xls", header=None)

    # 4 columns: acct, gap from ss:Index="3", name, amount. The column the
    # MergeAcross swallows is trailing and empty, so it is not materialised.
    assert df.shape == (2, 4)
    assert df.iloc[0].tolist()[0] == "acct"
    assert df.iloc[1, 0] == "111"
    assert df.iloc[1, 2] == "NH TMCP ĐT&PT Việt Nam"
    assert df.iloc[1, 3] == 210491420093.0


def _spreadsheetml_named(sheet_name: str) -> bytes:
    doc = (
        '<?xml version="1.0" encoding="UTF-16"?>'
        '<Workbook xmlns="urn:schemas-microsoft-com:office:spreadsheet"'
        ' xmlns:ss="urn:schemas-microsoft-com:office:spreadsheet">'
        f'<Worksheet ss:Name="{sheet_name}"><Table>'
        '<Row><Cell><Data ss:Type="String">v</Data></Cell></Row>'
        "</Table></Worksheet></Workbook>"
    )
    return ("﻿" + doc).encode("utf-16-be")


def _mock_xls(rm, name: str, content: bytes) -> None:
    rm.get(f"https://graph.microsoft.com/v1.0/me/drive/root:/{name}").mock(
        return_value=httpx.Response(
            200, json={"@microsoft.graph.downloadUrl": f"https://dl.url/{name}"}
        )
    )
    rm.get(f"https://dl.url/{name}").mock(
        return_value=httpx.Response(200, content=content)
    )


async def test_read_excel_selects_sheet_renamed_by_conversion(env_vars, mock_token):
    # xlsx cannot hold "Q1/Q2:[x]" as a title, so the conversion renames it.
    # Selecting by the name the source document uses must still work.
    with respx.mock(assert_all_called=False) as rm:
        _mock_xls(rm, "odd.xls", _spreadsheetml_named("Q1/Q2:[x]"))
        async with GraphClient("das_u1") as client:
            df = await client.read_excel(
                item_path="odd.xls", sheet="Q1/Q2:[x]", header=None
            )
    assert df.iat[0, 0] == "v"


async def test_read_excel_many_matches_renamed_sheet(env_vars, mock_token):
    with respx.mock(assert_all_called=False) as rm:
        _mock_xls(rm, "odd.xls", _spreadsheetml_named("Q1/Q2:[x]"))
        async with GraphClient("das_u1") as client:
            df, status = await client.read_excel_many(
                ["odd.xls"],
                sheet="q1/q2:[x]",
                sheet_match="ci",
                header=None,
                return_status=True,
            )
    assert status["status"].tolist() == ["success"]
    assert df.iat[0, 0] == "v"


async def test_read_excel_many_glob_pattern_is_not_sanitised(env_vars, mock_token):
    # "[12]" is a character class, not text to be replaced — a glob request
    # must reach match_sheet_name untouched.
    with respx.mock(assert_all_called=False) as rm:
        _mock_xls(rm, "g.xls", _spreadsheetml_named("Sales1"))
        async with GraphClient("das_u1") as client:
            df, status = await client.read_excel_many(
                ["g.xls"],
                sheet="Sales[12]",
                sheet_match="glob",
                header=None,
                return_status=True,
            )
    assert status["status"].tolist() == ["success"]
    assert df.iat[0, 0] == "v"


async def test_read_excel_many_handles_spreadsheetml(env_vars, mock_token):
    import pandas as pd

    with respx.mock(assert_all_called=False) as rm:
        for name in ("jan", "feb"):
            rm.get(
                f"https://graph.microsoft.com/v1.0/me/drive/root:/TB/{name}.xls"
            ).mock(
                return_value=httpx.Response(
                    200,
                    json={
                        "@microsoft.graph.downloadUrl": f"https://dl.url/{name}.xls"
                    },
                )
            )
            rm.get(f"https://dl.url/{name}.xls").mock(
                return_value=httpx.Response(200, content=_spreadsheetml_bytes())
            )
        async with GraphClient("das_u1") as client:
            df, status_df = await client.read_excel_many(
                ["TB/jan.xls", "TB/feb.xls"],
                sheet="Sheet1",
                header=None,
                return_status=True,
            )

    assert isinstance(df, pd.DataFrame)
    assert (status_df["status"] == "success").all()
    # 2 rows per file, 4 columns plus the _source column from add_source_column
    assert df.shape == (4, 5)
    # header=None, so each file contributes its header row plus its data row
    assert df[3].tolist() == [
        "amount",
        210491420093.0,
        "amount",
        210491420093.0,
    ]
    assert sorted(df["_source"].unique()) == ["TB/feb.xls", "TB/jan.xls"]


async def test_get_metadata_encodes_special_characters_in_path(env_vars, mock_token):
    # Registering the route at the encoded URL means the request only
    # matches if the client actually percent-encodes "?" before sending it —
    # unencoded, "?b.xlsx" would be parsed as a query string and this route
    # (and the real Graph endpoint) would never see it as part of the path.
    with respx.mock(assert_all_called=False) as rm:
        rm.get(
            "https://graph.microsoft.com/v1.0/me/drive/root:/Reports/a%3Fb.xlsx"
        ).mock(return_value=httpx.Response(200, json={"name": "a?b.xlsx", "size": 5}))

        async with GraphClient("das_u1") as client:
            meta = await client.get_metadata(item_path="Reports/a?b.xlsx")
    assert meta.name == "a?b.xlsx"


async def test_search_files_escapes_apostrophe(env_vars, mock_token):
    # OData escapes a literal "'" inside a string literal by doubling it.
    # Registering the route with the doubled-then-encoded literal proves the
    # fix: a plain percent-encode of "O'Brien" (O%27Brien) would not match.
    with respx.mock(assert_all_called=False) as rm:
        rm.get(
            "https://graph.microsoft.com/v1.0/me/drive/root/search(q='O%27%27Brien')"
            "?$top=25"
        ).mock(
            return_value=httpx.Response(
                200, json={"value": [{"name": "O'Brien notes.docx", "size": 1}]}
            )
        )
        async with GraphClient("das_u1") as client:
            results = await client.search_files("O'Brien")
    assert len(results) == 1
    assert results[0].name == "O'Brien notes.docx"


async def test_move_file_encodes_special_characters_in_dest(env_vars, mock_token):
    with respx.mock(assert_all_called=False) as rm:
        route = rm.patch("https://graph.microsoft.com/v1.0/me/drive/root:/a.xlsx")
        route.mock(return_value=httpx.Response(200, json={"name": "a.xlsx"}))

        async with GraphClient("das_u1") as client:
            await client.move_file(item_path="a.xlsx", dest_folder_path="Dest #1")

    payload = route.calls[0].request.read().decode()
    assert "Dest%20%231" in payload
    assert "Dest #1" not in payload
